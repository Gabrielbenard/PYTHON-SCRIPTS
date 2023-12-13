
#Quebrando o excel em abas na mesma planilha
from openpyxl import load_workbook
import os

nome_arquivo= "C:\\Users\\GBERNARDINO\\PycharmProjects\\AutomationRPA\\Quebrar.xlsx"

planilha_aberta = load_workbook(nome_arquivo)

sheet_chosen = planilha_aberta["Dados"]

totalLinha = len(sheet_chosen['A']) + 1

nomeNovo=""

for linha in range(2,totalLinha):
    nomeAtual = sheet_chosen["A%s" % linha].value

    if nomeNovo == nomeAtual:
        RowQuebra = len(sheet_chosen2["A"]) + 1
        celulaColunaA = "A" + str(RowQuebra)
        celulaColunaB = "B" + str(RowQuebra)
        celulaColunaC = "C" + str(RowQuebra)

        sheet_chosen2[celulaColunaA] = sheet_chosen["A%s" % RowQuebra].value
        sheet_chosen2[celulaColunaB] = sheet_chosen["B%s" % RowQuebra].value
        sheet_chosen2[celulaColunaC] = sheet_chosen["C%s" % RowQuebra].value

    else:
        #cria uma nova sheet com nome = nomeAtual
        sheet_resumo = planilha_aberta.create_sheet(title = nomeAtual)

        #Seleciona a Sheet criada
        sheet_chosen2 = planilha_aberta[nomeAtual]

        nomeNovo = sheet_chosen["A%s" % linha].value

        sheet_chosen2["A1"] = "Vendedor"
        sheet_chosen2["B1"] = "Produtos"
        sheet_chosen2["C1"] = "Vendas"
        sheet_chosen2["A2"] = sheet_chosen["A%s" % linha].value
        sheet_chosen2["B2"] = sheet_chosen["B%s" % linha].value
        sheet_chosen2["C2"] = sheet_chosen["C%s" % linha].value


planilha_aberta.save(nome_arquivo)


os.startfile((nome_arquivo))