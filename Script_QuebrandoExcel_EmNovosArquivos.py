from openpyxl import load_workbook
from openpyxl import Workbook
import os

nome_arquivo= "C:\\Users\\GBERNARDINO\\PycharmProjects\\AutomationRPA\\Quebrar.xlsx"

planilha_aberta = load_workbook(nome_arquivo)

sheet_chosen = planilha_aberta["Dados"]

creatingNewFileExcel = Workbook()

totalLinha = len(sheet_chosen['A']) + 1

nomeNovo=""

for linha in range(2,totalLinha):
    nomeAtual = sheet_chosen["A%s" % linha].value

    if nomeNovo == nomeAtual:
        RowQuebra = len(selectNewSpreedSheetVendas["A"]) + 1
        celulaColunaA = "A" + str(RowQuebra)
        celulaColunaB = "B" + str(RowQuebra)
        celulaColunaC = "C" + str(RowQuebra)

        selectNewSpreedSheetVendas[celulaColunaA] = sheet_chosen["A%s" % RowQuebra].value
        selectNewSpreedSheetVendas[celulaColunaB] = sheet_chosen["B%s" % RowQuebra].value
        selectNewSpreedSheetVendas[celulaColunaC] = sheet_chosen["C%s" % RowQuebra].value

    else:

        newSpreedSheet = creatingNewFileExcel.active
        newSpreedSheet.title = "Vendas"
        selectNewSpreedSheetVendas = creatingNewFileExcel["Vendas"]

        nomeNovo = sheet_chosen["A%s" % linha].value

        newExcelName = nomeNovo
        newExcelPath= f"C:\\Users\\GBERNARDINO\\PycharmProjects\\AutomationRPA\\Quebrando\\{newExcelName}.xlsx"

        selectNewSpreedSheetVendas["A1"] = "Vendedor"
        selectNewSpreedSheetVendas["B1"] = "Produtos"
        selectNewSpreedSheetVendas["C1"] = "Vendas"
        selectNewSpreedSheetVendas["A2"] = sheet_chosen["A%s" % linha].value
        selectNewSpreedSheetVendas["B2"] = sheet_chosen["B%s" % linha].value
        selectNewSpreedSheetVendas["C2"] = sheet_chosen["C%s" % linha].value

        selectNewSpreedSheetVendas.delete_rows(4, 100)

        creatingNewFileExcel.save(filename=newExcelPath)


os.startfile((nome_arquivo))